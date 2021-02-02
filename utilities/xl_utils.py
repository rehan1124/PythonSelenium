import openpyxl


def get_row_count(filename, sheetname):
    """
    Get number of rows from excel file
    :param filename:
    :type filename:
    :param sheetname:
    :type sheetname:
    :return:
    :rtype:
    """
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    return sheet.max_row


def get_column_count(filename, sheetname):
    """
    Get number of columns from excel file
    :param filename:
    :type filename:
    :param sheetname:
    :type sheetname:
    :return:
    :rtype:
    """
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    return sheet.max_column


def read_data(filename, sheetname, row_nm, column_nm):
    """
    Read data from given Row and Column
    :param filename:
    :type filename:
    :param sheetname:
    :type sheetname:
    :param row_nm:
    :type row_nm:
    :param column_nm:
    :type column_nm:
    :return:
    :rtype:
    """
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    return sheet.cell(row=row_nm, column=column_nm).value


def write_data(filename, sheetname, row_nm, column_nm, data):
    """
    Write data to a given Row and Column
    :param filename:
    :type filename:
    :param sheetname:
    :type sheetname:
    :param row_nm:
    :type row_nm:
    :param column_nm:
    :type column_nm:
    :param data:
    :type data:
    :return:
    :rtype:
    """
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    sheet.cell(row=row_nm, column=column_nm).value = data
    wb.save(filename)


def collect_all_data(filename, sheetname):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]

    data = []

    # for i in range(2, sheet.max_row):
    #     for j in range(1, sheet.max_column+1):
    #         data.append(sheet.cell(i, j).value)
    # for rows in sheet.iter_rows():
    #     data.append([cell.value for cell in rows])

    for rows in wb.value:
        data.append(rows)

    return data


if __name__ == "__main__":
    FILE_NAME = "../TestData/LoginData.xlsx"
    SHEET_NAME = "Sheet1"
    r = get_row_count(FILE_NAME, SHEET_NAME)
    c = get_column_count(FILE_NAME, SHEET_NAME)
    print(f"Number of rows and columns: {r}, {c}")

    sample_read = read_data(FILE_NAME, SHEET_NAME, 3, 1)
    print(sample_read)

    print(collect_all_data(FILE_NAME, SHEET_NAME))
